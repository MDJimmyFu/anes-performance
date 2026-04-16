#!/usr/bin/env python3
"""
import_excel.py
Extract monthly case data from the anesthesia performance Excel file
and write to data/cases/YYYY-MM.json files.

Usage:
    cd /Users/chunhsienfu/Documents/GitHub/anes-performance/scripts
    pip install openpyxl
    python import_excel.py
"""

import json
import os
import re
import uuid
import math
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

# ── Paths ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path(
    "/Users/chunhsienfu/Library/CloudStorage/"
    "OneDrive-個人/Anesthesiology/輔大麻醉科/輔麻績效試算.xlsx"
)
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "cases"
SETTINGS_OUT = Path(__file__).parent.parent / "data" / "point_settings.json"

# ── Skip these sheet names ─────────────────────────────────────────────────
SKIP_SHEETS = {"範本", "統計", "工作表1", "統計 (OLD)"}
SKIP_PREFIX = {"點數設定"}   # sheets starting with this prefix

# ── Point settings (4 periods, hard-coded from specification) ──────────────
POINT_SETTINGS = [
    {
        "id": "202209",
        "label": "202209",
        "effective_from": "2022-09",
        "effective_to": "2024-09",
        "methods": {
            "GE":       {"base": 1801.82, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "GM":       {"base": 1647.72, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "IV":       {"base": 735.08,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EA":       {"base": 1616.9,  "ot24": 117.76, "ot4plus": 117.76, "overtime": True},
            "SA":       {"base": 1102.16, "ot24": 107.64, "ot4plus": 107.64, "overtime": True},
            "Painless": {"base": 5105.8,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "Painless夜間": {"base": 5105.8, "ot24": 0,   "ot4plus": 0,      "overtime": False},
            "傳染GE":   {"base": 1356.586,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "困難氣道GE": {"base": 1176.588,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "HMC":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C/G":      {"base": 1029,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C+G":      {"base": 1715.196,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "ERCP":     {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EUS":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
        },
        "bonus_multipliers": {
            "無":1.0,"心臟手術":1.2,"腦部手術":1.2,"休克":1.2,"急診":1.2,
            "器官移植":2.0,"<6mo":2.0,"6mo-2yo":1.8,"2yo-7yo":1.6,"自費麻醉":1.3,"醫美":1.5
        },
        "extras": {
            "GVL_AWS_MAC":382.2,"Rusch_Video":588.0,"OMT":1401.792,"A_line":271.124,
            "CVC":473.34,"PAC":885.5,"TEE":1545.6,"CO":92.0,"Optiflow":1200.5,
            "BIS_self":288.12,"BIS_NHI_adult":605.36,"BIS_NHI_child":968.576,
            "blanket":367.5,"IVPCA":1137.29,"NBPCA":2806.524,"PCEA":2771.44,
            "PCA_days":514.5,"IV_sedation":8575.0,"ultrasound":483.0,"ByBIS":2001.405
        }
    },
    {
        "id": "202410",
        "label": "202410",
        "effective_from": "2024-10",
        "effective_to": "2025-05",
        "methods": {
            "GE":       {"base": 1801.82, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "GM":       {"base": 1647.72, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "IV":       {"base": 735.08,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EA":       {"base": 1616.9,  "ot24": 117.76, "ot4plus": 117.76, "overtime": True},
            "SA":       {"base": 1102.16, "ot24": 107.64, "ot4plus": 107.64, "overtime": True},
            "Painless": {"base": 5105.8,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "Painless夜間": {"base": 5105.8,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "傳染GE":   {"base": 1356.586,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "困難氣道GE": {"base": 1176.588,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "HMC":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C/G":      {"base": 1029,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C+G":      {"base": 1715.196,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "ERCP":     {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EUS":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
        },
        "bonus_multipliers": {
            "無":1.0,"心臟手術":1.2,"腦部手術":1.2,"休克":1.2,"急診":1.2,
            "器官移植":2.0,"<6mo":2.0,"6mo-2yo":1.8,"2yo-7yo":1.6,"自費麻醉":1.3,"醫美":1.5
        },
        "extras": {
            "GVL_AWS_MAC":382.2,"Rusch_Video":588.0,"OMT":1401.792,"A_line":271.124,
            "CVC":473.34,"PAC":885.5,"TEE":1545.6,"CO":92.0,"Optiflow":1200.5,
            "BIS_self":288.12,"BIS_NHI_adult":605.36,"BIS_NHI_child":968.576,
            "blanket":367.5,"IVPCA":1137.29,"NBPCA":2664.277,"PCEA":2771.44,
            "PCA_days":514.5,"IV_sedation":8575.0,"ultrasound":483.0,"ByBIS":2001.405
        }
    },
    {
        "id": "202506",
        "label": "202506",
        "effective_from": "2025-06",
        "effective_to": "2025-08",
        "methods": {
            "GE":       {"base": 1801.82, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "GM":       {"base": 1647.72, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "IV":       {"base": 735.08,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EA":       {"base": 1616.9,  "ot24": 117.76, "ot4plus": 117.76, "overtime": True},
            "SA":       {"base": 1102.16, "ot24": 107.64, "ot4plus": 107.64, "overtime": True},
            "Painless": {"base": 5105.8,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "Painless夜間": {"base": 5105.8,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "傳染GE":   {"base": 1356.586,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "困難氣道GE": {"base": 1176.588,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "HMC":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C/G":      {"base": 1029,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C+G":      {"base": 1715.196,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "ERCP":     {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EUS":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
        },
        "bonus_multipliers": {
            "無":1.0,"心臟手術":1.2,"腦部手術":1.2,"休克":1.2,"急診":1.2,
            "器官移植":2.0,"<6mo":2.0,"6mo-2yo":1.8,"2yo-7yo":1.6,"自費麻醉":1.3,"醫美":1.5
        },
        "extras": {
            "GVL_AWS_MAC":382.2,"Rusch_Video":588.0,"OMT":1401.792,"A_line":271.124,
            "CVC":473.34,"PAC":885.5,"TEE":1545.6,"CO":92.0,"Optiflow":1200.5,
            "BIS_self":288.12,"BIS_NHI_adult":605.36,"BIS_NHI_child":968.576,
            "blanket":360.5,"IVPCA":1137.29,"NBPCA":2394.777,"PCEA":2771.44,
            "PCA_days":514.5,"IV_sedation":8575.0,"ultrasound":483.0,"ByBIS":2001.405
        }
    },
    {
        "id": "202509",
        "label": "202509",
        "effective_from": "2025-09",
        "effective_to": None,
        "methods": {
            "GE":       {"base": 1801.82, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "GM":       {"base": 1647.72, "ot24": 411.7,  "ot4plus": 514.74, "overtime": True},
            "IV":       {"base": 735.08,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EA":       {"base": 1616.9,  "ot24": 117.76, "ot4plus": 117.76, "overtime": True},
            "SA":       {"base": 1102.16, "ot24": 107.64, "ot4plus": 107.64, "overtime": True},
            "Painless": {"base": 5105.8,  "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "Painless夜間": {"base": 5105.8,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "傳染GE":   {"base": 1356.586,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "困難氣道GE": {"base": 1176.588,"ot24": 0,    "ot4plus": 0,      "overtime": False},
            "HMC":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C/G":      {"base": 1029,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "C+G":      {"base": 1715.196,"ot24": 0,      "ot4plus": 0,      "overtime": False},
            "ERCP":     {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
            "EUS":      {"base": 1715,    "ot24": 0,      "ot4plus": 0,      "overtime": False},
        },
        "bonus_multipliers": {
            "無":1.0,"心臟手術":1.2,"腦部手術":1.2,"休克":1.2,"急診":1.2,
            "器官移植":2.0,"<6mo":2.0,"6mo-2yo":1.8,"2yo-7yo":1.6,"自費麻醉":1.3,"醫美":1.5
        },
        "extras": {
            "GVL_AWS_MAC":382.2,"Rusch_Video":588.0,"OMT":1401.792,"A_line":271.124,
            "CVC":473.34,"PAC":885.5,"TEE":1545.6,"CO":92.0,"Optiflow":1200.5,
            "BIS_self":288.12,"BIS_NHI_adult":605.36,"BIS_NHI_child":968.576,
            "blanket":360.5,"IVPCA":1137.29,"NBPCA":2394.777,"PCEA":2771.44,
            "PCA_days":514.5,"IV_sedation":8575.0,"ultrasound":483.0,"ByBIS":2001.405
        }
    },
]


# ── Helpers ────────────────────────────────────────────────────────────────

def safe_float(val, default=0.0):
    """Convert cell value to float safely."""
    if val is None:
        return default
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    return int(safe_float(val, default))


def safe_str(val, default=""):
    if val is None:
        return default
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "nan") else default


def sheet_name_to_ym(name):
    """
    Convert sheet name like "2026.03" or "2026-03" or "202603" to "2026-03".
    Returns None if not parseable as a year-month.
    """
    name = name.strip()
    # "2026.03"
    m = re.match(r'^(\d{4})[.\-/](\d{1,2})$', name)
    if m:
        y, mo = m.group(1), m.group(2).zfill(2)
        return f"{y}-{mo}"
    # "202603"
    m = re.match(r'^(\d{4})(\d{2})$', name)
    if m:
        y, mo = m.group(1), m.group(2)
        return f"{y}-{mo}"
    return None


def get_applicable_settings(year_month):
    """Find the point settings period for a given YYYY-MM string."""
    sorted_periods = sorted(POINT_SETTINGS, key=lambda p: p["effective_from"], reverse=True)
    for p in sorted_periods:
        frm = p["effective_from"]
        to  = p["effective_to"]
        if year_month >= frm and (to is None or year_month <= to):
            return p
    return sorted_periods[-1]  # fallback to earliest


def calculate_base_performance(method, duration, settings):
    methods = settings.get("methods", {})
    m = methods.get(method)
    if not m:
        return 0.0
    base = m["base"]
    if m.get("overtime") and duration > 120:
        units  = int((duration - 120) // 30)
        first4 = min(units, 4)
        beyond = max(units - 4, 0)
        return base + first4 * m["ot24"] + beyond * m["ot4plus"]
    return base


def calculate_extras(row_dict, settings):
    ex = settings.get("extras", {})
    total = 0.0
    for key in ["GVL_AWS_MAC","Rusch_Video","OMT","A_line","CVC","PAC","TEE","CO",
                "Optiflow","BIS_self","BIS_NHI_adult","BIS_NHI_child","blanket",
                "IVPCA","NBPCA","PCEA","PCA_days","IV_sedation","ultrasound","ByBIS"]:
        total += row_dict.get(key, 0) * ex.get(key, 0)
    return total


def calculate_total(case_dict, settings):
    method   = case_dict.get("method", "")
    duration = case_dict.get("duration", 0)
    bonus    = case_dict.get("bonus", "無")
    handover = case_dict.get("handover", 1.0)

    base   = calculate_base_performance(method, duration, settings)
    mult   = settings.get("bonus_multipliers", {}).get(bonus, 1.0)
    before = base * mult * handover
    extras = calculate_extras(case_dict, settings)
    return round(before + extras, 3)


# ── Column index constants (0-based) ──────────────────────────────────────
COL_DATE        = 0
COL_CASE_NO     = 1
COL_DIAGNOSIS   = 2
COL_ASA         = 3
COL_BONUS       = 4
COL_METHOD      = 5
COL_POINTS      = 6   # already-calculated column (we recalculate anyway)
COL_HANDOVER    = 7
COL_DURATION    = 8
COL_GVL         = 9
COL_RUSCH       = 10
COL_OMT         = 11
COL_ALINE       = 12
COL_CVC         = 13
COL_PAC         = 14
COL_TEE         = 15
COL_CO          = 16
COL_OPTIFLOW    = 17
COL_BIS_SELF    = 18
COL_BIS_NHI_A   = 19
COL_BIS_NHI_C   = 20
COL_BLANKET     = 21
COL_IVPCA       = 22
COL_NBPCA       = 23
COL_PCEA        = 24
COL_PCA_DAYS    = 25
COL_IV_SED      = 26
COL_ULTRASOUND  = 27
COL_BYBIS       = 28


def parse_date(val, year_month):
    """Try to extract a YYYY-MM-DD date from the cell."""
    if val is None:
        return None
    from datetime import datetime, date as date_t
    if isinstance(val, (datetime, date_t)):
        if isinstance(val, datetime):
            return val.date().isoformat()
        return val.isoformat()
    s = str(val).strip()
    # "2026/3/4" or "2026-03-04"
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    # day-only integer
    try:
        day = int(float(s))
        if 1 <= day <= 31:
            y, m = year_month.split("-")
            return f"{y}-{m}-{str(day).zfill(2)}"
    except (ValueError, TypeError):
        pass
    return None


def parse_handover(val):
    """Map handover value to 1, 0.8, or 0.2."""
    f = safe_float(val, 1.0)
    if abs(f - 0.8) < 0.01:
        return 0.8
    if abs(f - 0.2) < 0.01:
        return 0.2
    return 1.0


def parse_asa(val):
    """Return ASA as string e.g. '2' or '3E'."""
    s = safe_str(val, "2")
    # might be float like 2.0
    try:
        s = str(int(float(s)))
    except (ValueError, TypeError):
        pass
    return s if s else "2"


VALID_METHODS = {
    "GE","GM","IV","EA","SA","Painless","Painless夜間",
    "傳染GE","困難氣道GE","HMC","C/G","C+G","ERCP","EUS"
}

VALID_BONUSES = {
    "無","心臟手術","腦部手術","休克","急診","器官移植",
    "<6mo","6mo-2yo","2yo-7yo","自費麻醉","醫美"
}


def parse_row(row_vals, year_month):
    """
    row_vals: list of cell values (already 0-indexed, starting from col 0).
    Returns a case dict or None if the row should be skipped.
    """
    def v(idx):
        return row_vals[idx] if idx < len(row_vals) else None

    method = safe_str(v(COL_METHOD))
    if not method or method not in VALID_METHODS:
        return None  # skip rows with invalid/missing method

    date = parse_date(v(COL_DATE), year_month)
    if not date:
        # Try to construct from year_month
        date = year_month + "-01"

    case_no   = safe_str(v(COL_CASE_NO))
    diagnosis = safe_str(v(COL_DIAGNOSIS))
    asa       = parse_asa(v(COL_ASA))
    bonus_raw = safe_str(v(COL_BONUS), "無")
    bonus     = bonus_raw if bonus_raw in VALID_BONUSES else "無"
    handover  = parse_handover(v(COL_HANDOVER))
    duration  = safe_int(v(COL_DURATION), 0)

    extras = {
        "GVL_AWS_MAC":   safe_int(v(COL_GVL)),
        "Rusch_Video":   safe_int(v(COL_RUSCH)),
        "OMT":           safe_int(v(COL_OMT)),
        "A_line":        safe_int(v(COL_ALINE)),
        "CVC":           safe_int(v(COL_CVC)),
        "PAC":           safe_int(v(COL_PAC)),
        "TEE":           safe_int(v(COL_TEE)),
        "CO":            safe_int(v(COL_CO)),
        "Optiflow":      safe_int(v(COL_OPTIFLOW)),
        "BIS_self":      safe_int(v(COL_BIS_SELF)),
        "BIS_NHI_adult": safe_int(v(COL_BIS_NHI_A)),
        "BIS_NHI_child": safe_int(v(COL_BIS_NHI_C)),
        "blanket":       safe_int(v(COL_BLANKET)),
        "IVPCA":         safe_int(v(COL_IVPCA)),
        "NBPCA":         safe_int(v(COL_NBPCA)),
        "PCEA":          safe_int(v(COL_PCEA)),
        "PCA_days":      safe_int(v(COL_PCA_DAYS)),
        "IV_sedation":   safe_int(v(COL_IV_SED)),
        "ultrasound":    safe_int(v(COL_ULTRASOUND)),
        "ByBIS":         safe_int(v(COL_BYBIS)),
    }

    settings   = get_applicable_settings(year_month)
    base_pts   = calculate_base_performance(method, duration, settings)
    case_data  = {"method": method, "duration": duration, "bonus": bonus,
                  "handover": handover, **extras, "date": date}
    total_pts  = calculate_total(case_data, settings)

    return {
        "id":               str(uuid.uuid4()),
        "date":             date,
        "case_no":          case_no,
        "diagnosis":        diagnosis,
        "asa":              asa,
        "bonus":            bonus,
        "method":           method,
        "base_points":      round(base_pts, 3),
        "handover":         handover,
        "duration":         duration,
        **extras,
        "total_performance": total_pts,
        "notes":            "",
    }


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at:\n  {EXCEL_PATH}")
        print("Please update EXCEL_PATH in this script.")
        return

    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    total_cases = 0
    month_counts = {}

    for sheet_name in wb.sheetnames:
        # Skip unwanted sheets
        if sheet_name in SKIP_SHEETS:
            print(f"  SKIP (blacklist): {sheet_name}")
            continue
        if any(sheet_name.startswith(p) for p in SKIP_PREFIX):
            print(f"  SKIP (prefix): {sheet_name}")
            continue

        year_month = sheet_name_to_ym(sheet_name)
        if not year_month:
            print(f"  SKIP (not a month): {sheet_name}")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Data starts at row index 2 (0-based), i.e. Excel row 3
        data_rows = rows[2:] if len(rows) > 2 else []

        cases = []
        for row in data_rows:
            row_list = list(row)
            case = parse_row(row_list, year_month)
            if case:
                cases.append(case)

        out_path = OUTPUT_DIR / f"{year_month}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(cases, f, ensure_ascii=False, indent=2)

        month_counts[year_month] = len(cases)
        total_cases += len(cases)
        print(f"  {sheet_name} → {year_month}.json ({len(cases)} cases)")

    # Write point settings
    with open(SETTINGS_OUT, "w", encoding="utf-8") as f:
        json.dump(POINT_SETTINGS, f, ensure_ascii=False, indent=2)
    print(f"\nWrote: {SETTINGS_OUT}")

    print(f"\nDone! Total cases: {total_cases}")
    print(f"Month breakdown:")
    for ym in sorted(month_counts):
        print(f"  {ym}: {month_counts[ym]}")


if __name__ == "__main__":
    main()
